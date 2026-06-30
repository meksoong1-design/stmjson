
# ============================================================
# STM Document AI Bank Statement Parser v6.1
# Document AI API + Smart SCB Column Parser + Regex Fallback
# แก้ปัญหา:
# - SCB อ่านยอดหลักพันหาย
# - footer/header หลุดเป็นรายการ
# - KeyError: bank_name จาก session เก่า
# - Export Excel ครบชีท
# ============================================================

import os
import re
import tempfile
from datetime import datetime

import pandas as pd
import streamlit as st

from google.api_core.client_options import ClientOptions
from google.cloud import documentai_v1 as documentai
from google.oauth2 import service_account
from google.protobuf.json_format import MessageToDict

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, Border, Side, PatternFill


# ============================================================
# 0) PAGE CONFIG & CONSTANTS
# ============================================================

st.set_page_config(page_title="STM Document AI Parser v6.1", layout="centered")

st.markdown("""
<style>
.block-container {
    padding-top: 2rem;
    padding-bottom: 3rem;
    max-width: 980px;
}
h1 {
    font-size: 2.35rem !important;
    font-weight: 800 !important;
}
.section-title {
    display:flex;
    align-items:center;
    gap:8px;
    font-weight:800;
    font-size:1.15rem;
    margin:18px 0 10px 0;
}
.section-title .step-number {
    display:inline-flex;
    align-items:center;
    justify-content:center;
    min-width:30px;
    height:30px;
    border-radius:999px;
    background:#4c82fb;
    color:white;
    font-weight:800;
}
div[data-testid="stButton"] button,
div[data-testid="stDownloadButton"] button {
    border-radius:10px;
    font-weight:750;
}
div[data-testid="stFileUploader"] section {
    border-radius:10px;
}
.small-note {
    font-size: 0.85rem;
    color: #666;
}
</style>
""", unsafe_allow_html=True)

BALANCE_TOLERANCE = 0.05
BANK_LABELS = {
    "AUTO": "ตรวจจับอัตโนมัติ",
    "KBANK": "กสิกรไทย (KBANK)",
    "BBL": "กรุงเทพ (BBL)",
    "SCB": "ไทยพาณิชย์ (SCB)",
    "KRUNGSRI": "กรุงศรี (Krungsri)",
    "DEFAULT": "ไม่ระบุ / อื่นๆ",
}

EMPTY_COLS = [
    "ลำดับ",
    "วันที่เดือนปี",
    "เวลา",
    "รายการ",
    "เดบิต",
    "เครดิต",
    "ยอดคงเหลือ",
    "รายละเอียด",
    "ช่องทาง",
    "ไฟล์ต้นฉบับ",
]


# ============================================================
# 1) SESSION / LOGIN
# ============================================================

if "authenticated" not in st.session_state:
    st.session_state.authenticated = False

if "results" not in st.session_state:
    st.session_state.results = None


def login_page():
    st.title("Login")
    st.caption("กรุณาเข้าสู่ระบบก่อนใช้งาน STM Document AI Parser")

    password = st.text_input("Password", type="password")

    if st.button("เข้าสู่ระบบ", type="primary", use_container_width=True):
        correct_pw = st.secrets.get("APP_PASSWORD", "stm2025")

        if password == correct_pw:
            st.session_state.authenticated = True
            st.session_state.results = None
            st.rerun()
        else:
            st.error("รหัสผ่านไม่ถูกต้อง")

    st.stop()


# ============================================================
# 2) DOCUMENT AI API
# ============================================================

def get_documentai_client():
    if "gcp_service_account" not in st.secrets:
        st.error("ไม่พบ [gcp_service_account] ใน Streamlit Secrets")
        st.stop()

    try:
        info = dict(st.secrets["gcp_service_account"])

        if "private_key" in info and isinstance(info["private_key"], str):
            info["private_key"] = info["private_key"].replace("\\n", "\n")

        creds = service_account.Credentials.from_service_account_info(
            info,
            scopes=["https://www.googleapis.com/auth/cloud-platform"],
        )

        project_id = str(st.secrets["DOCUMENTAI_PROJECT_ID"])
        location = str(st.secrets["DOCUMENTAI_LOCATION"])
        processor_id = str(st.secrets["DOCUMENTAI_PROCESSOR_ID"])

        opts = ClientOptions(api_endpoint=f"{location}-documentai.googleapis.com")
        client = documentai.DocumentProcessorServiceClient(
            credentials=creds,
            client_options=opts,
        )

        processor_name = client.processor_path(project_id, location, processor_id)

        return client, processor_name

    except Exception as e:
        st.error(f"การเชื่อมต่อ API ไม่สำเร็จ: {e}")
        st.stop()


def get_mime_type(filename):
    name = filename.lower()

    if name.endswith(".pdf"):
        return "application/pdf"

    if name.endswith((".jpg", ".jpeg")):
        return "image/jpeg"

    if name.endswith(".png"):
        return "image/png"

    raise ValueError("รองรับเฉพาะ PDF, JPG, JPEG, PNG")


@st.cache_data(show_spinner=False)
def call_document_ai(file_bytes, filename):
    client, processor_name = get_documentai_client()

    raw_doc = documentai.RawDocument(
        content=file_bytes,
        mime_type=get_mime_type(filename),
    )

    req = documentai.ProcessRequest(
        name=processor_name,
        raw_document=raw_doc,
    )

    result = client.process_document(request=req)
    document = result.document

    doc_dict = MessageToDict(
        document._pb,
        preserving_proto_field_name=True,
    )

    raw_text = document.text or ""

    return doc_dict, raw_text


# ============================================================
# 3) TEXT / DATE / MONEY HELPERS
# ============================================================

def clean_text(value):
    if value is None:
        return ""

    return re.sub(
        r"\s+",
        " ",
        str(value).replace("\x00", " ").strip(),
    )


def normalize_ocr_text(text):
    text = clean_text(text)
    text = text.replace("O", "0").replace("o", "0")
    text = text.replace("I", "1").replace("l", "1")
    return text


def parse_money(s):
    if s is None:
        return None

    s = normalize_ocr_text(str(s))
    s = re.sub(r"\s+", "", s)

    # 1.234,56
    if re.match(r"^-?\d{1,3}(\.\d{3})+,\d{2}$", s):
        s = s.replace(".", "").replace(",", ".")

    # 1234,56
    elif re.match(r"^-?\d+,\d{2}$", s):
        s = s.replace(",", ".")

    # 1,234.56
    else:
        s = s.replace(",", "")

    s = re.sub(r"[^0-9.\-]", "", s)

    if not s:
        return None

    try:
        return float(s)
    except Exception:
        return None


def extract_money_values(text):
    text = normalize_ocr_text(str(text))

    patterns = [
        r"-?\d{1,3}(?:,\d{3})+\.\d{2}",
        r"-?\d{1,3}(?:,\d{3})+,\d{2}",
        r"-?\d+\.\d{2}",
        r"-?\d+,\d{2}",
    ]

    vals = []

    for pat in patterns:
        for m in re.findall(pat, text):
            v = parse_money(m)

            if v is not None:
                vals.append(v)

    return vals


def extract_money_values_v2(text):
    text = normalize_ocr_text(str(text))

    pattern = (
        r"-?\d{1,3}(?:,\d{3})+\.\d{2}"
        r"|-?\d{1,3}(?:,\d{3})+,\d{2}"
        r"|-?\d+\.\d{2}"
        r"|-?\d+,\d{2}"
    )

    vals = []

    for m in re.findall(pattern, text):
        s = clean_text(m)

        # กรณี OCR เป็น 1,010,36 ให้ถือว่า comma ตัวสุดท้ายคือทศนิยม
        if re.match(r"^-?\d{1,3}(?:,\d{3})+,\d{2}$", s):
            head, dec = s.rsplit(",", 1)
            s = head.replace(",", "") + "." + dec
        else:
            s = s.replace(",", "")

        try:
            vals.append(float(s))
        except Exception:
            pass

    return vals


def extract_date(text):
    text = normalize_ocr_text(str(text))
    text = re.sub(r"[.]", "-", text)

    m = re.search(r"\b(\d{1,2})[-/](\d{1,2})[-/](\d{2,4})\b", text)

    if m:
        d = int(m.group(1))
        mo = int(m.group(2))
        y = int(m.group(3))

        year = 2000 + y if y < 100 else y

        try:
            return datetime(year, mo, d).strftime("%d/%m/%Y")
        except Exception:
            return None

    return None


def extract_time(text):
    m = re.search(r"\b\d{1,2}:\d{2}(?::\d{2})?\b", str(text))
    return m.group(0) if m else ""


def parse_date_for_sort(d):
    try:
        return datetime.strptime(str(d), "%d/%m/%Y")
    except Exception:
        return datetime.max


def parse_time_for_sort(t):
    try:
        return datetime.strptime(str(t), "%H:%M").time()
    except Exception:
        try:
            return datetime.strptime(str(t), "%H:%M:%S").time()
        except Exception:
            return datetime.max.time()


def safe_float(value, default=0.0):
    try:
        if value is None:
            return default
        return float(value)
    except Exception:
        return default


# ============================================================
# 4) ACCOUNT INFO
# ============================================================

def extract_account_info(raw_text):
    owner = "ไม่ระบุ"
    acc_no = "ไม่ระบุ"

    lines = [clean_text(x) for x in raw_text.splitlines() if clean_text(x)]

    for i, line in enumerate(lines[:90]):
        if owner == "ไม่ระบุ":
            if re.search(r"(ชื่อบัญชี|Account Name|ชื่อ\s*-\s*นามสกุล|ชื่อ\s*สกุล|Name)", line, re.I):
                owner_candidate = re.sub(
                    r"^.*?(ชื่อบัญชี|Account\s*Name|ชื่อ\s*-\s*นามสกุล|ชื่อ\s*สกุล|Name)\s*[:：]?\s*",
                    "",
                    line,
                    flags=re.I,
                ).strip()

                if owner_candidate and not re.search(r"Address|Account|Date|เลขที่", owner_candidate, re.I):
                    owner = owner_candidate
                elif i + 1 < len(lines):
                    owner = lines[i + 1]

            elif re.match(r"^(นาย|นาง|น\.ส\.|นางสาว|MR\.|MRS\.|MS\.)\s+", line, re.I):
                owner = line

        if acc_no == "ไม่ระบุ":
            m = re.search(
                r"(\d{3}[- ]?\d{3}[- ]?\d{1,4}[- ]?\d?|\d{3}[- ]?\d{1}[- ]?\d{5}[- ]?\d{1}|\d{10,12})",
                line,
            )

            if m:
                acc_no = m.group(1).replace(" ", "")

    return owner, acc_no


# ============================================================
# 5) DOCUMENT AI TOKEN HELPERS
# ============================================================

def token_text_from_anchor(layout, full_text):
    segs = (
        layout.get("textAnchor", {}).get("textSegments", [])
        or layout.get("text_anchor", {}).get("text_segments", [])
    )

    parts = []

    for seg in segs:
        s = int(seg.get("startIndex", seg.get("start_index", 0)) or 0)
        e = int(seg.get("endIndex", seg.get("end_index", 0)) or 0)
        parts.append(full_text[s:e])

    return clean_text("".join(parts))


def get_doc_tokens(doc_dict, full_text):
    tokens = []

    for page_no, page in enumerate(doc_dict.get("pages", []), start=1):
        for token in page.get("tokens", []):
            layout = token.get("layout", {})

            verts = (
                layout.get("boundingPoly", {}).get("normalizedVertices", [])
                or layout.get("bounding_poly", {}).get("normalized_vertices", [])
            )

            if not verts:
                continue

            xs = [v.get("x", 0) for v in verts]
            ys = [v.get("y", 0) for v in verts]

            text = token_text_from_anchor(layout, full_text)

            if not text:
                continue

            tokens.append({
                "page": page_no,
                "text": text,
                "x": sum(xs) / len(xs),
                "y": sum(ys) / len(ys),
                "x_min": min(xs),
                "x_max": max(xs),
                "y_min": min(ys),
                "y_max": max(ys),
            })

    return tokens


def join_tokens(tokens, money_mode=False):
    tokens = sorted(tokens, key=lambda t: t["x"])

    if money_mode:
        return "".join(t["text"] for t in tokens)

    return clean_text(" ".join(t["text"] for t in tokens))


def rebuild_lines_from_json(doc_dict, full_text):
    pages = doc_dict.get("pages", [])
    all_lines = []

    for page in pages:
        tokens_info = []

        for token in page.get("tokens", []):
            layout = token.get("layout", {})

            verts = (
                layout.get("boundingPoly", {}).get("normalizedVertices", [])
                or layout.get("bounding_poly", {}).get("normalized_vertices", [])
            )

            if not verts:
                continue

            x_center = sum([v.get("x", 0) for v in verts]) / len(verts)
            y_center = sum([v.get("y", 0) for v in verts]) / len(verts)

            token_text = token_text_from_anchor(layout, full_text)

            if token_text:
                tokens_info.append({
                    "text": token_text,
                    "x": x_center,
                    "y": y_center,
                })

        tokens_info = sorted(tokens_info, key=lambda k: k["y"])

        current_line = []
        current_y = None

        for tok in tokens_info:
            if current_y is None:
                current_line.append(tok)
                current_y = tok["y"]

            elif abs(tok["y"] - current_y) < 0.01:
                current_line.append(tok)

            else:
                current_line = sorted(current_line, key=lambda k: k["x"])
                line_text = " ".join([t["text"] for t in current_line])
                all_lines.append(line_text)

                current_line = [tok]
                current_y = tok["y"]

        if current_line:
            current_line = sorted(current_line, key=lambda k: k["x"])
            line_text = " ".join([t["text"] for t in current_line])
            all_lines.append(line_text)

    return all_lines


# ============================================================
# 6) SCB SPECIAL PARSER
# ============================================================

def is_scb_statement(raw_text):
    u = raw_text.upper()

    return (
        "SIAM COMMERCIAL BANK" in u
        or "SCB" in u
        or "ไทยพาณิชย์" in raw_text
        or "STATEMENT OF SAVING ACCOUNT" in u
    )


def scb_direction(row_text):
    u = row_text.upper()

    # SCB จาก statement:
    # X1 = Credit / เงินเข้า
    # X2 = Debit / เงินออก
    # SIPI = Debit
    if re.search(r"\bX1\b", u):
        return "credit"

    if re.search(r"\bX2\b", u) or re.search(r"\bSIPI\b", u):
        return "debit"

    return ""


def parse_scb_transactions_from_json(doc_dict, full_text):
    tokens = get_doc_tokens(doc_dict, full_text)

    rows = []

    # ค่าคอลัมน์ SCB จาก layout statement แนวนอน A4
    COL_DATE = (0.020, 0.095)
    COL_TIME = (0.075, 0.145)
    COL_CODE = (0.120, 0.190)
    COL_CHANNEL = (0.170, 0.270)
    COL_AMOUNT = (0.300, 0.520)
    COL_BALANCE = (0.520, 0.680)
    COL_DESC = (0.660, 0.985)

    y_tolerance = 0.008

    date_anchors = []

    for t in tokens:
        if not (COL_DATE[0] <= t["x"] <= COL_DATE[1]):
            continue

        if extract_date(t["text"]):
            date_anchors.append(t)

    date_anchors = sorted(date_anchors, key=lambda t: (t["page"], t["y"]))

    for anchor in date_anchors:
        page = anchor["page"]
        y = anchor["y"]

        row_tokens = [
            t for t in tokens
            if t["page"] == page and abs(t["y"] - y) <= y_tolerance
        ]

        row_text = join_tokens(row_tokens)
        direction = scb_direction(row_text)

        if direction not in ["debit", "credit"]:
            continue

        date_text = join_tokens([
            t for t in row_tokens
            if COL_DATE[0] <= t["x"] <= COL_DATE[1]
        ])

        time_text = join_tokens([
            t for t in row_tokens
            if COL_TIME[0] <= t["x"] <= COL_TIME[1]
        ])

        code_text = join_tokens([
            t for t in row_tokens
            if COL_CODE[0] <= t["x"] <= COL_CODE[1]
        ])

        channel_text = join_tokens([
            t for t in row_tokens
            if COL_CHANNEL[0] <= t["x"] <= COL_CHANNEL[1]
        ])

        amount_text = join_tokens(
            [
                t for t in row_tokens
                if COL_AMOUNT[0] <= t["x"] <= COL_AMOUNT[1]
            ],
            money_mode=True,
        )

        balance_text = join_tokens(
            [
                t for t in row_tokens
                if COL_BALANCE[0] <= t["x"] <= COL_BALANCE[1]
            ],
            money_mode=True,
        )

        desc_text = join_tokens([
            t for t in row_tokens
            if COL_DESC[0] <= t["x"] <= COL_DESC[1]
        ])

        date = extract_date(date_text) or extract_date(row_text)
        time = extract_time(time_text) or extract_time(row_text)

        amount_vals = extract_money_values_v2(amount_text)
        balance_vals = extract_money_values_v2(balance_text)

        amount = amount_vals[-1] if amount_vals else 0.0
        balance = balance_vals[-1] if balance_vals else 0.0

        debit = amount if direction == "debit" else 0.0
        credit = amount if direction == "credit" else 0.0

        if not date or not time:
            continue

        if amount <= 0 and balance <= 0:
            continue

        rows.append({
            "วันที่เดือนปี": date,
            "เวลา": time,
            "รายการ": clean_text(f"{code_text} {channel_text}"),
            "รายละเอียด": desc_text,
            "เดบิต": debit,
            "เครดิต": credit,
            "ยอดคงเหลือ": balance,
            "ช่องทาง": channel_text or "SCB",
            "_page": page,
            "_y": y,
        })

    df = pd.DataFrame(rows)

    if df.empty:
        return df

    df = df.sort_values(["_page", "_y"], kind="stable").reset_index(drop=True)
    df = df.drop(columns=["_page", "_y"], errors="ignore")

    return df


# ============================================================
# 7) FALLBACK REGEX PARSER
# ============================================================

def parse_transactions(lines, bank_choice):
    rows = []
    prev_balance = None

    for line in lines:
        date = extract_date(line)
        time = extract_time(line)
        moneys = extract_money_values(line)

        desc = line

        for m in re.findall(
            r"-?\d{1,3}(?:,\d{3})+\.\d{2}|-?\d+\.\d{2}|\b\d{1,2}:\d{2}(?::\d{2})?\b|\b\d{1,2}[-/]\d{1,2}[-/]\d{2,4}\b",
            line,
        ):
            desc = desc.replace(m, "")

        desc = clean_text(desc)

        if any(kw in line.upper() for kw in ["ยอดยกมา", "BROUGHT FORWARD", "B/F"]):
            if moneys:
                prev_balance = moneys[-1]
            continue

        if not date:
            continue

        debit = 0.0
        credit = 0.0
        balance = 0.0

        if len(moneys) >= 2:
            amount = moneys[-2]
            balance = moneys[-1]

            if prev_balance is not None:
                if abs(prev_balance - amount - balance) < BALANCE_TOLERANCE:
                    debit = amount

                elif abs(prev_balance + amount - balance) < BALANCE_TOLERANCE:
                    credit = amount

                else:
                    upper_line = line.upper()

                    if any(kw in upper_line for kw in ["X1", "ฝาก", "รับ", "DEPOSIT", "CREDIT"]):
                        credit = amount

                    elif any(kw in upper_line for kw in ["X2", "ถอน", "จ่าย", "WITHDRAWAL", "DEBIT"]):
                        debit = amount

                    elif balance < prev_balance:
                        debit = amount

                    else:
                        credit = amount

            else:
                upper_line = line.upper()

                if any(kw in upper_line for kw in ["X1", "ฝาก", "รับ", "DEPOSIT", "CREDIT"]):
                    credit = amount

                elif any(kw in upper_line for kw in ["X2", "ถอน", "จ่าย", "WITHDRAWAL", "DEBIT"]):
                    debit = amount

                else:
                    debit = amount

            prev_balance = balance

        elif len(moneys) == 1:
            balance = moneys[0]
            prev_balance = balance

        rows.append({
            "วันที่เดือนปี": date,
            "เวลา": time,
            "รายการ": desc[:60],
            "รายละเอียด": desc,
            "เดบิต": debit,
            "เครดิต": credit,
            "ยอดคงเหลือ": balance,
            "ช่องทาง": "Document AI",
        })

    return pd.DataFrame(rows)


# ============================================================
# 8) CLEAN TRANSACTION DF
# ============================================================

def clean_transaction_df(df):
    if df is None or df.empty:
        return pd.DataFrame(columns=EMPTY_COLS)

    df = df.copy()

    for col in ["วันที่เดือนปี", "เวลา", "รายการ", "รายละเอียด", "ช่องทาง", "ไฟล์ต้นฉบับ"]:
        if col not in df.columns:
            df[col] = ""

    for col in ["เดบิต", "เครดิต", "ยอดคงเหลือ"]:
        if col not in df.columns:
            df[col] = 0.0

        df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0.0)

    df = df[df["วันที่เดือนปี"].astype(str).str.contains(r"\d{2}/\d{2}/\d{4}", regex=True, na=False)]
    df = df[df["เวลา"].astype(str).str.contains(r"\d{1,2}:\d{2}", regex=True, na=False)]

    bad_kw = [
        "AUTO-GENERATED",
        "SIGNATURE",
        "DOCUMENT",
        "PAGE",
        "หน้า",
        "TOTAL AMOUNTS",
        "TOTAL ITEMS",
        "BALANCE BROUGHT FORWARD",
        "ยอดเงินคงเหลือยกมา",
        "STATEMENT OF",
        "ACCOUNT NO",
        "DATE",
    ]

    bad_pattern = "|".join([re.escape(x) for x in bad_kw])

    combined_text = (
        df["รายละเอียด"].astype(str)
        + " "
        + df["รายการ"].astype(str)
    ).str.upper()

    df = df[~combined_text.str.contains(bad_pattern, regex=True, na=False)]

    money_sum = (
        pd.to_numeric(df["เดบิต"], errors="coerce").fillna(0)
        + pd.to_numeric(df["เครดิต"], errors="coerce").fillna(0)
        + pd.to_numeric(df["ยอดคงเหลือ"], errors="coerce").fillna(0)
    )

    df = df[money_sum > 0]

    df["_sort_date"] = df["วันที่เดือนปี"].apply(parse_date_for_sort)
    df["_sort_time"] = df["เวลา"].apply(parse_time_for_sort)
    df["_row_order"] = range(len(df))

    df = df.sort_values(
        ["_sort_date", "_sort_time", "_row_order"],
        kind="stable",
    ).drop(columns=["_sort_date", "_sort_time", "_row_order"], errors="ignore")

    df = df.reset_index(drop=True)

    if "ลำดับ" in df.columns:
        df = df.drop(columns=["ลำดับ"], errors="ignore")

    df.insert(0, "ลำดับ", range(1, len(df) + 1))

    for col in EMPTY_COLS:
        if col not in df.columns:
            df[col] = ""

    return df[EMPTY_COLS]


# ============================================================
# 9) SALARY / BONUS DETECTION
# ============================================================

def detect_salary(df_txn):
    if df_txn.empty:
        return pd.DataFrame()

    salary_kw = [
        "SALARY",
        "PAYROLL",
        "SAL",
        "PAYR",
        "เงินเดือน",
        "FAST(SAL)",
    ]

    bonus_kw = [
        "BONUS",
        "INCENTIVE",
        "COMMISSION",
        "โบนัส",
        "อินเซนทีฟ",
        "ค่าคอม",
    ]

    rows = []

    for _, row in df_txn.iterrows():
        cr = safe_float(row.get("เครดิต", 0))

        if cr <= 0:
            continue

        desc = str(row.get("รายละเอียด", "")).upper()

        if any(k.upper() in desc for k in bonus_kw):
            rows.append({
                "กลุ่ม": "รายได้พิเศษ/โบนัส",
                "วันที่": row["วันที่เดือนปี"],
                "จำนวนเงิน": cr,
            })

        elif any(k.upper() in desc for k in salary_kw):
            rows.append({
                "กลุ่ม": "เงินเดือน",
                "วันที่": row["วันที่เดือนปี"],
                "จำนวนเงิน": cr,
            })

    return pd.DataFrame(rows)


def build_summary_df(df_txn, owner, acc_no, bank_name):
    if df_txn.empty:
        last_bal = 0.0
        s_date = ""
        e_date = ""
    else:
        valid = pd.to_numeric(df_txn["ยอดคงเหลือ"], errors="coerce").dropna()
        last_bal = float(valid.iloc[-1]) if len(valid) else 0.0
        s_date = df_txn["วันที่เดือนปี"].iloc[0]
        e_date = df_txn["วันที่เดือนปี"].iloc[-1]

    debit_sum = float(df_txn["เดบิต"].sum()) if not df_txn.empty else 0.0
    credit_sum = float(df_txn["เครดิต"].sum()) if not df_txn.empty else 0.0

    rows = [
        {"รายการ": "เจ้าของบัญชี", "ข้อมูล": owner},
        {"รายการ": "เลขที่บัญชีเงินฝาก", "ข้อมูล": acc_no},
        {"รายการ": "ธนาคาร", "ข้อมูล": bank_name},
        {"รายการ": "จำนวนรายการ", "ข้อมูล": len(df_txn)},
        {"รายการ": "รวมเดบิต", "ข้อมูล": debit_sum},
        {"รายการ": "รวมเครดิต", "ข้อมูล": credit_sum},
        {"รายการ": "ยอดคงเหลือล่าสุด", "ข้อมูล": last_bal},
        {"รายการ": "วันที่เริ่มต้น", "ข้อมูล": s_date},
        {"รายการ": "วันที่สิ้นสุด", "ข้อมูล": e_date},
    ]

    df_income = detect_salary(df_txn)

    for grp in ["เงินเดือน", "รายได้พิเศษ/โบนัส"]:
        rows.append({"รายการ": "หมวดหมู่รายได้", "ข้อมูล": grp})

        if not df_income.empty and grp in df_income["กลุ่ม"].values:
            g = df_income[df_income["กลุ่ม"] == grp].copy()
            total = float(g["จำนวนเงิน"].sum())

            rows += [
                {"รายการ": f"{grp} - สถานะ", "ข้อมูล": "พบ"},
                {"รายการ": f"{grp} - จำนวนรายการ", "ข้อมูล": len(g)},
                {"รายการ": f"{grp} - ยอดรวม", "ข้อมูล": total},
                {"รายการ": f"{grp} - ค่าเฉลี่ยต่อเดือน", "ข้อมูล": total / len(g)},
            ]

            for i, r in enumerate(g.itertuples(), 1):
                rows += [
                    {"รายการ": f"{grp} - รายการที่ {i}", "ข้อมูล": r.วันที่},
                    {"รายการ": f"{grp} - ยอด", "ข้อมูล": r.จำนวนเงิน},
                ]

        else:
            rows += [
                {"รายการ": f"{grp} - สถานะ", "ข้อมูล": "ไม่พบ"},
                {"รายการ": f"{grp} - จำนวนรายการ", "ข้อมูล": 0},
                {"รายการ": f"{grp} - ยอดรวม", "ข้อมูล": 0.0},
            ]

        rows.append({"รายการ": "", "ข้อมูล": ""})

    rows.append({
        "รายการ": "สร้างไฟล์เมื่อ",
        "ข้อมูล": datetime.now().strftime("%d/%m/%Y %H:%M:%S"),
    })

    return pd.DataFrame(rows)


# ============================================================
# 10) BANK STATEMENT 1 SHEET
# ============================================================

def calculate_30day_blocks(df_txn, period_days=30, periods=3):
    empty = {
        "daily": pd.DataFrame(columns=["วันที่", "ยอดคงเหลือสิ้นวัน", "ช่วงที่"]),
        "blocks": [],
    }

    if df_txn.empty:
        return empty

    work = df_txn.copy()

    work["_date"] = pd.to_datetime(
        work["วันที่เดือนปี"].astype(str).str.replace("-", "/"),
        dayfirst=True,
        errors="coerce",
    )

    work = work.dropna(subset=["_date"])

    if work.empty:
        return empty

    work["ยอดคงเหลือ"] = pd.to_numeric(
        work["ยอดคงเหลือ"],
        errors="coerce",
    ).fillna(0.0)

    work = work.reset_index(drop=True)
    work["_row_order"] = work.index
    work = work.sort_values(["_date", "_row_order"])
    work["_day"] = work["_date"].dt.normalize()

    daily_txn = work.groupby("_day").agg(
        ยอดคงเหลือสิ้นวัน=("ยอดคงเหลือ", "last"),
    ).sort_index()

    full_days = pd.date_range(
        start=daily_txn.index.min(),
        end=daily_txn.index.max(),
        freq="D",
    )

    daily = daily_txn.reindex(full_days)
    daily["ยอดคงเหลือสิ้นวัน"] = daily["ยอดคงเหลือสิ้นวัน"].ffill().fillna(0.0)

    daily = daily.tail(period_days * periods).reset_index().rename(
        columns={"index": "วันที่"},
    )

    daily["ลำดับ"] = range(1, len(daily) + 1)
    daily["ช่วงที่"] = ((daily["ลำดับ"] - 1) // period_days) + 1

    blocks = []

    for b in range(1, periods + 1):
        blk = daily[daily["ช่วงที่"] == b].copy()

        blocks.append({
            "ช่วงที่": b,
            "จำนวนวัน": len(blk),
            "ผลรวม": float(blk["ยอดคงเหลือสิ้นวัน"].sum()) if len(blk) else 0.0,
            "ค่าเฉลี่ย": float(blk["ยอดคงเหลือสิ้นวัน"].mean()) if len(blk) else 0.0,
            "data": blk,
        })

    return {
        "daily": daily,
        "blocks": blocks,
    }


def add_bank_statement_sheet(path, df_txn, owner, acc_no, bank_name):
    result = calculate_30day_blocks(df_txn)
    blocks = result["blocks"]

    wb = load_workbook(path)
    ws = wb.create_sheet("BANK STATEMENT 1")

    title_blue = PatternFill("solid", fgColor="6FA1E8")
    header_blue = PatternFill("solid", fgColor="A9C4F5")
    light_blue = PatternFill("solid", fgColor="DDEBF7")
    yellow = PatternFill("solid", fgColor="FFFF00")
    gray = PatternFill("solid", fgColor="D9D9D9")

    heavy = Border(
        left=Side(style="medium", color="000000"),
        right=Side(style="medium", color="000000"),
        top=Side(style="medium", color="000000"),
        bottom=Side(style="medium", color="000000"),
    )

    thin = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9"),
    )

    ws.merge_cells("A1:I1")
    ws["A1"] = "BANK STATEMENT 1"
    ws["A1"].font = Font(bold=True, size=13)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws["A1"].fill = title_blue

    avg_refs = ["C37", "F37", "I37"]

    for idx in range(3):
        r = 2 + idx
        ws.cell(r, 1, 1)
        ws.cell(r, 2, f"เดือนที่ {idx + 1}")
        ws.cell(r, 3, f"=ROUND({avg_refs[idx]},2)")

    ws.cell(2, 5, "เฉลี่ยต่อเดือน")
    ws.cell(
        2,
        6,
        "=IF((C39>0)+(F39>0)+(I39>0)=0,0,ROUND((IF(C39>0,C37,0)+IF(F39>0,F37,0)+IF(I39>0,I37,0))/((C39>0)+(F39>0)+(I39>0)),2))",
    )

    ws.cell(3, 5, "จำนวนวันรวม")
    ws.cell(3, 6, "=SUM(C39,F39,I39)")

    ws.cell(4, 5, "ประเมินวงเงินได้")
    ws.cell(4, 6, "=ROUND(F2*G3*G4,2)")

    ws.merge_cells("H2:I2")
    ws["H2"] = "Account Detail"

    ws.cell(3, 8, "ธนาคาร")
    ws.cell(3, 9, "เลขที่บัญชี")

    ws.cell(4, 8, bank_name)
    ws.cell(4, 9, acc_no)

    for r in range(2, 5):
        for c in range(1, 10):
            cell = ws.cell(r, c)
            cell.fill = header_blue
            cell.border = heavy
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")

    for addr in ["C2", "C3", "C4", "F2", "F3"]:
        ws[addr].fill = light_blue

    for idx, block in enumerate(blocks):
        sc = [1, 4, 7][idx]
        nc, dc, ac = sc, sc + 1, sc + 2

        for col, val in [(nc, "N"), (dc, "DATE"), (ac, "AMOUNT")]:
            cell = ws.cell(5, col, val)
            cell.fill = header_blue
            cell.font = Font(bold=True)
            cell.border = heavy
            cell.alignment = Alignment(horizontal="center", vertical="center")

        data = block.get("data", pd.DataFrame())

        for i in range(1, 31):
            er = 6 + i - 1

            ws.cell(er, nc, i).fill = light_blue
            ws.cell(er, dc).fill = light_blue

            if i <= len(data):
                rd = data.iloc[i - 1]
                dv = rd.get("วันที่", "")
                av = float(rd.get("ยอดคงเหลือสิ้นวัน", 0))

                ws.cell(er, dc, dv)
                ws.cell(er, ac, av)

                if i > 1:
                    pv = float(data.iloc[i - 2].get("ยอดคงเหลือสิ้นวัน", 0))

                    if round(av, 2) != round(pv, 2):
                        ws.cell(er, dc).fill = yellow
                        ws.cell(er, ac).fill = yellow

            for col in [nc, dc, ac]:
                cell = ws.cell(er, col)
                cell.border = thin
                cell.alignment = Alignment(
                    horizontal="center" if col != ac else "right",
                    vertical="center",
                )

            ws.cell(er, dc).number_format = "d mmm yy"
            ws.cell(er, ac).number_format = "#,##0.00"

        amount_letter = ws.cell(5, ac).column_letter

        ws.cell(36, nc, "รวม")
        ws.cell(36, ac, f"=SUM({amount_letter}6:{amount_letter}35)")
        ws.cell(37, ac, f"=IF({amount_letter}39=0,0,{amount_letter}36/{amount_letter}39)")
        ws.cell(38, ac, "จำนวนวัน")
        ws.cell(39, ac, f"=COUNT({amount_letter}6:{amount_letter}35)")

        for r in range(36, 40):
            for c in [nc, dc, ac]:
                cell = ws.cell(r, c)
                cell.border = heavy
                cell.alignment = Alignment(
                    horizontal="center" if c != ac else "right",
                    vertical="center",
                )

                if c == ac:
                    cell.number_format = "#,##0.00"

        ws.cell(38, ac).fill = gray
        ws.cell(38, ac).font = Font(bold=True)

    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A5"

    widths = {
        "A": 5,
        "B": 14,
        "C": 16,
        "D": 5,
        "E": 14,
        "F": 16,
        "G": 5,
        "H": 14,
        "I": 16,
    }

    for letter, width in widths.items():
        ws.column_dimensions[letter].width = width

    wb.save(path)


# ============================================================
# 11) EXCEL FORMAT
# ============================================================

def auto_fit_worksheet(ws):
    ws.freeze_panes = "A2"

    thin = Side(style="thin", color="D9D9D9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    def estimate_text_width(value):
        w = 0.0

        for ch in clean_text(value):
            if "\u0e00" <= ch <= "\u0e7f":
                w += 1.4
            elif ch.isupper():
                w += 1.1
            elif ch.isdigit() or ch in ".,:-/()":
                w += 0.85
            elif ch == " ":
                w += 0.4
            else:
                w += 0.95

        return w

    headers = {
        c: clean_text(ws.cell(row=1, column=c).value)
        for c in range(1, ws.max_column + 1)
    }

    for row in ws.iter_rows():
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical="center", wrap_text=False)

            if cell.row == 1:
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center", vertical="center")

            if isinstance(cell.value, (int, float)):
                cell.number_format = "#,##0.00"
                cell.alignment = Alignment(horizontal="right", vertical="center")

    for col_cells in ws.columns:
        letter = col_cells[0].column_letter
        col_idx = col_cells[0].column

        width = estimate_text_width(headers.get(col_idx, ""))

        for cell in col_cells:
            if cell.value is not None:
                width = max(width, estimate_text_width(str(cell.value)))

        ws.column_dimensions[letter].width = max(8, min(width + 1.5, 45))


def format_summary_sheet(ws):
    ws.sheet_view.showGridLines = False

    h_fill = PatternFill("solid", fgColor="1F4E79")
    s_fill = PatternFill("solid", fgColor="DDEBF7")
    sub_fill = PatternFill("solid", fgColor="EAF4EA")

    for cell in ws[1]:
        cell.fill = h_fill
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for r in range(2, ws.max_row + 1):
        label = clean_text(ws.cell(r, 1).value)

        if label == "หมวดหมู่รายได้":
            for c in range(1, 3):
                ws.cell(r, c).fill = s_fill
                ws.cell(r, c).font = Font(bold=True)

        if any(k in label for k in ["สถานะ", "จำนวนรายการ", "ยอดรวม", "ค่าเฉลี่ยต่อเดือน"]):
            ws.cell(r, 1).fill = sub_fill
            ws.cell(r, 2).fill = sub_fill
            ws.cell(r, 1).font = Font(bold=True)

        if isinstance(ws.cell(r, 2).value, (int, float)):
            ws.cell(r, 2).number_format = "#,##0.00"


def build_output_excel(df_account, df_txn, owner, acc_no, bank_name):
    with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
        tmp_path = tmp.name

    with pd.ExcelWriter(tmp_path, engine="openpyxl") as writer:
        df_account.to_excel(writer, sheet_name="เจ้าของบัญชี", index=False)
        df_txn.to_excel(writer, sheet_name="รายละเอียด เดบิต-เครดิต", index=False)
        build_summary_df(df_txn, owner, acc_no, bank_name).to_excel(
            writer,
            sheet_name="สรุปยอด",
            index=False,
        )

    add_bank_statement_sheet(tmp_path, df_txn, owner, acc_no, bank_name)

    wb = load_workbook(tmp_path)

    for ws in wb.worksheets:
        if ws.title != "BANK STATEMENT 1":
            auto_fit_worksheet(ws)

            if ws.title == "สรุปยอด":
                format_summary_sheet(ws)

    wb.save(tmp_path)

    with open(tmp_path, "rb") as f:
        data = f.read()

    os.unlink(tmp_path)

    return data


# ============================================================
# 12) MAIN FILE PROCESSOR
# ============================================================

def detect_bank_name(raw_text, selected_bank="AUTO"):
    if selected_bank and selected_bank != "AUTO":
        return selected_bank

    if is_scb_statement(raw_text):
        return "SCB"

    u = raw_text.upper()

    if "KASIKORNBANK" in u or "KBANK" in u or "กสิกร" in raw_text:
        return "KBANK"

    if "BANGKOK BANK" in u or "BBL" in u or "กรุงเทพ" in raw_text:
        return "BBL"

    if "KRUNGSRI" in u or "AYUDHYA" in u or "กรุงศรี" in raw_text:
        return "KRUNGSRI"

    return "AUTO"


def process_one_file(file_bytes, filename, selected_bank):
    doc_dict, raw_text = call_document_ai(file_bytes, filename)

    if not raw_text.strip():
        raise ValueError("Document AI อ่านข้อความไม่เจอ กรุณาตรวจสอบไฟล์")

    owner, acc_no = extract_account_info(raw_text)
    bank_name = detect_bank_name(raw_text, selected_bank)

    if bank_name == "SCB":
        df_txn = parse_scb_transactions_from_json(doc_dict, raw_text)

        # ถ้า SCB parser อ่านไม่ได้จริง ๆ ให้ fallback
        if df_txn.empty:
            lines = rebuild_lines_from_json(doc_dict, raw_text)
            df_txn = parse_transactions(lines, selected_bank)
    else:
        lines = rebuild_lines_from_json(doc_dict, raw_text)
        df_txn = parse_transactions(lines, selected_bank)

    if df_txn is None or df_txn.empty:
        df_txn = pd.DataFrame(columns=[
            "วันที่เดือนปี",
            "เวลา",
            "รายการ",
            "รายละเอียด",
            "เดบิต",
            "เครดิต",
            "ยอดคงเหลือ",
            "ช่องทาง",
        ])

    df_txn["ไฟล์ต้นฉบับ"] = filename
    df_txn = clean_transaction_df(df_txn)

    return {
        "df_txn": df_txn,
        "owner": owner,
        "acc_no": acc_no,
        "bank_name": bank_name,
    }


# ============================================================
# 13) STREAMLIT UI
# ============================================================

if not st.session_state.authenticated:
    login_page()

# ปุ่มนี้แก้ปัญหา session เก่าที่เคยมี key ไม่ครบ เช่น bank_name
with st.sidebar:
    st.markdown("### เครื่องมือ")
    if st.button("Reset session / ล้างข้อมูลค้าง", use_container_width=True):
        st.session_state.results = None
        st.rerun()

st.title("STM Document AI Parser v6.1")
st.caption("เชื่อม API | อ่าน SCB ตามคอลัมน์จริง | แก้ยอดหลักพันหาย | ตัด footer/header | Export Excel ครบชีท")
st.divider()

st.markdown(
    '<div class="section-title"><span class="step-number">1</span><span>ตรวจสอบ / แก้ไขข้อมูลบัญชี</span></div>',
    unsafe_allow_html=True,
)

st.info("ระบบจะพยายามค้นหา ชื่อ และ เลขบัญชี ให้อัตโนมัติจากไฟล์ Statement")

col1, col2 = st.columns(2)

owner_input = col1.text_input(
    "ชื่อเจ้าของบัญชี",
    placeholder="เว้นว่างให้ระบบหาอัตโนมัติ",
)

acc_input = col2.text_input(
    "เลขที่บัญชี",
    placeholder="เว้นว่างให้ระบบหาอัตโนมัติ",
)

st.markdown(
    '<div class="section-title"><span class="step-number">2</span><span>เลือกธนาคาร / อัปโหลด Statement</span></div>',
    unsafe_allow_html=True,
)

selected_bank = st.selectbox(
    "เลือกธนาคาร",
    options=list(BANK_LABELS.keys()),
    format_func=lambda x: BANK_LABELS[x],
    index=0,
)

uploaded_files = st.file_uploader(
    "เลือกไฟล์ Statement เพื่อส่งให้ API ประมวลผล",
    type=["pdf", "jpg", "jpeg", "png"],
    accept_multiple_files=True,
)

process_clicked = st.button(
    "เริ่มประมวลผลผ่าน API",
    type="primary",
    use_container_width=True,
    disabled=not uploaded_files,
)

if process_clicked:
    # กันผลลัพธ์เก่าค้างก่อนเริ่มรอบใหม่
    st.session_state.results = None

    progress = st.progress(0, text="เริ่มส่งไฟล์ไปที่ Document AI...")
    status = st.empty()

    try:
        all_txn = []
        auto_owner = ""
        auto_acc = ""
        detected_bank = "AUTO"

        total = len(uploaded_files)

        for idx, uf in enumerate(uploaded_files, start=1):
            status.info(f"API กำลังประมวลผลไฟล์ {idx}/{total}: {uf.name}")
            progress.progress(int((idx - 1) / total * 70))

            res = process_one_file(
                uf.read(),
                uf.name,
                selected_bank,
            )

            if not res["df_txn"].empty:
                all_txn.append(res["df_txn"])

            if not auto_owner or auto_owner == "ไม่ระบุ":
                if res.get("owner", "ไม่ระบุ") != "ไม่ระบุ":
                    auto_owner = res.get("owner", "ไม่ระบุ")

            if not auto_acc or auto_acc == "ไม่ระบุ":
                if res.get("acc_no", "ไม่ระบุ") != "ไม่ระบุ":
                    auto_acc = res.get("acc_no", "ไม่ระบุ")

            if detected_bank == "AUTO" and res.get("bank_name", "AUTO") != "AUTO":
                detected_bank = res.get("bank_name", "AUTO")

        progress.progress(80, text="กำลังจัดเรียงข้อมูลและสร้าง Excel...")

        if all_txn:
            final_df = pd.concat(all_txn, ignore_index=True)
        else:
            final_df = pd.DataFrame(columns=EMPTY_COLS)

        final_df = clean_transaction_df(final_df)

        owner = owner_input.strip() or auto_owner or "ไม่ระบุ"
        acc_no = acc_input.strip() or auto_acc or "ไม่ระบุ"
        bank_name = detected_bank if detected_bank != "AUTO" else selected_bank

        df_account = pd.DataFrame([{
            "เจ้าของบัญชี": owner,
            "เลขที่บัญชีเงินฝาก": acc_no,
            "ธนาคาร": bank_name,
        }])

        excel_bytes = build_output_excel(
            df_account,
            final_df,
            owner,
            acc_no,
            bank_name,
        )

        safe_owner = re.sub(r"[\\/:\*\?\"<>\|]", "_", owner)
        excel_name = f"Statement_{safe_owner}.xlsx"

        st.session_state.results = {
            "df": final_df,
            "excel": excel_bytes,
            "excel_name": excel_name,
            "owner": owner,
            "acc_no": acc_no,
            "bank_name": bank_name,
        }

        progress.progress(100, text="เสร็จสิ้น")
        status.success(f"ประมวลผลสำเร็จ พบ {len(final_df):,} รายการ")

    except Exception as e:
        progress.empty()
        status.error(f"เกิดข้อผิดพลาด: {e}")
        st.exception(e)


# ============================================================
# 14) RESULT DISPLAY
# ============================================================

if st.session_state.get("results"):
    res = st.session_state.results

    # กัน KeyError จาก session เก่า
    if not isinstance(res, dict):
        st.session_state.results = None
        st.rerun()

    if "df" not in res:
        st.session_state.results = None
        st.rerun()

    res.setdefault("bank_name", "AUTO")
    res.setdefault("owner", "ไม่ระบุ")
    res.setdefault("acc_no", "ไม่ระบุ")
    res.setdefault("excel_name", "Statement.xlsx")

    df = res["df"]

    if df is None:
        df = pd.DataFrame(columns=EMPTY_COLS)

    st.divider()

    debit_sum = float(
        pd.to_numeric(
            df.get("เดบิต", pd.Series(dtype=float)),
            errors="coerce",
        ).fillna(0).sum()
    )

    credit_sum = float(
        pd.to_numeric(
            df.get("เครดิต", pd.Series(dtype=float)),
            errors="coerce",
        ).fillna(0).sum()
    )

    last_balance = 0.0

    if not df.empty and "ยอดคงเหลือ" in df.columns:
        valid_bal = pd.to_numeric(df["ยอดคงเหลือ"], errors="coerce").dropna()

        if len(valid_bal):
            last_balance = float(valid_bal.iloc[-1])

    m1, m2, m3, m4 = st.columns(4)

    m1.metric("เจ้าของบัญชี", res.get("owner", "ไม่ระบุ"))
    m2.metric("จำนวนรายการ", f"{len(df):,}")
    m3.metric("รวมเดบิต", f"{debit_sum:,.2f}")
    m4.metric("รวมเครดิต", f"{credit_sum:,.2f}")

    m5, m6 = st.columns(2)
    m5.metric("ยอดคงเหลือล่าสุด", f"{last_balance:,.2f}")
    m6.metric("ธนาคาร", res.get("bank_name", "AUTO"))

    if "excel" in res and res["excel"]:
        st.download_button(
            "ดาวน์โหลดไฟล์ Excel",
            data=res["excel"],
            file_name=res.get("excel_name", "Statement.xlsx"),
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
            type="primary",
        )
    else:
        st.warning("ยังไม่มีไฟล์ Excel สำหรับดาวน์โหลด กรุณาประมวลผลใหม่อีกครั้ง")

    st.markdown("#### ตัวอย่างข้อมูลที่ประมวลผลได้")

    if df.empty:
        st.warning("ไม่พบรายการธุรกรรม")
    else:
        st.dataframe(
            df.head(80).style.format({
                "เดบิต": "{:,.2f}",
                "เครดิต": "{:,.2f}",
                "ยอดคงเหลือ": "{:,.2f}",
            }),
            use_container_width=True,
            height=500,
        )

    with st.expander("ตรวจสอบยอดรวม"):
        st.write({
            "จำนวนรายการ": int(len(df)),
            "รวมเดบิต": round(debit_sum, 2),
            "รวมเครดิต": round(credit_sum, 2),
            "ยอดคงเหลือล่าสุด": round(last_balance, 2),
            "ธนาคาร": res.get("bank_name", "AUTO"),
        })

    if st.button("ล้างข้อมูล", use_container_width=True):
        st.session_state.results = None
        st.rerun()
